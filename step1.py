import openpyxl

# Give the location of the file
path = "mapping.xlsx"


wb_obj = openpyxl.load_workbook(path)
the_map = []

sheets = ['SOFP', 'SOPL', 'SOCI', 'SOCF', 'SOCIE',
          'Subclassifications', 'Analysis of Income and Expense']


def ts(s):
    return([s[0], s[1:]])


count = 0
for sheet in sheets:
    sheet_obj = wb_obj.get_sheet_by_name(name=sheet)
    the_map.append([])

    for x in range(2, sheet_obj.max_row+1):
        template_cell = sheet_obj.cell(row=x, column=1)
        SOFP_cell = sheet_obj.cell(row=x, column=3)
        SOPL_cell = sheet_obj.cell(row=x, column=4)
        SOCI_cell = sheet_obj.cell(row=x, column=5)
        SOCF_cell = sheet_obj.cell(row=x, column=6)
        SOCIE_cell = sheet_obj.cell(row=x, column=7)
        Management_Accounts_cell = sheet_obj.cell(row=x, column=8)

        if(template_cell.value != None):

            uf = []
            tl = []

            if(SOFP_cell.value != None):
                for y in SOFP_cell.value.split(","):
                    el = ts(y)
                    tl.append(el)

            uf.append(tl)
            tl = []

            if(SOPL_cell.value != None):

                for y in SOPL_cell.value.split(","):
                    el = ts(y)
                    tl.append(el)
            uf.append(tl)
            tl = []

            if(SOCI_cell.value != None):

                for y in SOCI_cell.value.split(","):
                    el = ts(y)
                    tl.append(el)
            uf.append(tl)
            tl = []

            if(SOCF_cell.value != None):

                for y in SOCF_cell.value.split(","):
                    el = ts(y)
                    tl.append(el)
            uf.append(tl)
            tl = []

            if(SOCIE_cell.value != None):

                for y in SOCIE_cell.value.split(","):
                    el = ts(y)
                    tl.append(el)
            uf.append(tl)
            tl = []

            if(Management_Accounts_cell.value != None):

                for y in Management_Accounts_cell.value.split(","):
                    el = ts(y)
                    tl.append(el)
            uf.append(tl)
            tl = []

            the_map[count].append([template_cell.value, uf])

    count += 1

sheet_count = 0
for x in the_map:
    print("--------------", sheets[sheet_count], "--------------")
    for y in x:
        print(y)
    sheet_count += 1
    print("--------------SWITCHING", "--------------")


print("----------------------------------------------------OPERATION CHANGE----------------------------------------------------------")

hm_obj = openpyxl.load_workbook("fs.xlsx")

r_sheets = ["SOFP", "SOPL", "SOCI", "SOCF", "SOCIE"]

number_map = []

# r_sheet = hm_obj.get_sheet_by_name(name=r_sheets[2])
# for sh in the_map:
#     print("-----------------------------SHEET CHANGE------------------------------")
#     for l in sh:
#         m = 0
#         for k in l[1][2]:
#             print(r_sheet[k[1]].value)
#             try:
#                 if(k[0] == '+'):
#                     m = m+int(r_sheet[k[1]].value)
#                 if(k[0] == '-'):
#                     m = m-int(r_sheet[k[1]].value)
#             except:
#                 print('Funny Data')

#         l[1][0] = m


for x in range(len(r_sheets)):
    r_sheet = hm_obj.get_sheet_by_name(name=r_sheets[x])
    print("-----------------------------", r_sheet,
          "------------------------------")
    for sh in the_map:

        for l in sh:
            m = 0
            for k in l[1][x]:
                try:
                    if(k[0] == '+'):
                        m = m+int(r_sheet[k[1]].value)
                    if(k[0] == '-'):
                        m = m-int(r_sheet[k[1]].value)
                except:
                    print('Error loading FS workbook')

            l[1][x] = m


r_sheet = openpyxl.load_workbook("management_accounts.xlsx", data_only=True).get_active_sheet()
for sh in the_map:
    print("-----------------------------SHEET CHANGE------------------------------")
    for l in sh:
        m = 0
        for k in l[1][5]:
            try:
                if(k[0] == '+'):
                    m = m+int(r_sheet[k[1]].value)
                if(k[0] == '-'):
                    m = m-int(r_sheet[k[1]].value)
            except:
                print('Error loading Management Accounts workbook')
                print(r_sheet[k[1]].value)

        l[1][5] = m


print("----------------------------------------------------OPERATION CHANGE----------------------------------------------------------")

sheet_count = 0
for x in the_map:
    print("--------------", sheets[sheet_count], "--------------")
    for y in x:
        print(y)
    sheet_count += 1
    print("--------------SWITCHING", "--------------")

print("----------------------------------------------------OPERATION CHANGE----------------------------------------------------------")


sheet_count = 0
total = 0
for x in the_map:
    print("--------------", sheets[sheet_count], "--------------")
    for y in x:
        y[1] = sum(y[1])
        print(y)
    sheet_count += 1
    print("--------------SWITCHING", "--------------")


print("----------------------------------------------------OPERATION CHANGE----------------------------------------------------------")


wpath = 'template.xlsx'
wt = openpyxl.load_workbook(wpath)

sheets = ['SOFP', 'SOPL', 'SOCI', 'SOCF', 'SOCIE',
          'Subclassfications- A,L,E', 'Analysis of Income and Expense']


sheet_count = 0
for x in the_map:
    print("--------------", sheets[sheet_count], "--------------")
    so = wt.get_sheet_by_name(name=sheets[sheet_count])
    for y in x:
        print("At ", y[0])
        print("Previous Data - ", so[y[0]].value)
        so[y[0]].value = y[1]
        print("New Data - ", so[y[0]].value)
    sheet_count += 1
    print("--------------SWITCHING", "--------------")


wt.save(wpath)
wt.close

print(" Operation Successful")
print(" Developed by virapy22(2020)")