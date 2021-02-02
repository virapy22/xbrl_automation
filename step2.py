import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

# Give the location of the file
path = "xbrl_form_mapping.xlsx"


wb_obj = openpyxl.load_workbook(path)
the_map = []

sheets = ['StaOfFinPosCurNonCur', 'IncoStatByNatuOfExpe', 'SOfCIOciCPNetOfTax', 'StaOfCashFlowsIndMet',
          'StatementOfChangesInEquity', 'NotesSuOfAsLiEqCuNonCu', 'NotesAnaOfIncExpByNat']
sheets2 = ['SOFP', 'SOPL', 'SOCI', 'SOCF', 'SOCIE',
           'Subclassfications- A,L,E', 'Analysis of Income and Expense']


def ts(s):
    return([s[0], s[1:]])


redFill = PatternFill(start_color='FF0000',
                      end_color='FF0000',
                      fill_type='solid')

greenFill = PatternFill(start_color='00FF00',
                        end_color='00FF00',
                        fill_type='solid')

count = 0
for sheet in sheets:
    sheet_obj = wb_obj.get_sheet_by_name(name=sheet)
    the_map.append([])

    # cell_obj = sheet_obj.cell(row=1, column=1)
    # cell_obj1 = sheet_obj.cell(row=1, column=3)
    # the_map[count].append([cell_obj.value, cell_obj1.value])
    for x in range(2, sheet_obj.max_row+1):
        cell_obj = sheet_obj.cell(row=x, column=1)
        cell_obj1 = sheet_obj.cell(row=x, column=3)

        if(cell_obj1.value != None):
            the_map[count].append([cell_obj.value, ts(cell_obj1.value)])
    count += 1

for x in the_map:
    for y in x:
        print(y)

co_obj1 = openpyxl.load_workbook('xbrl_form.xlsx')
co_obj2 = openpyxl.load_workbook('template.xlsx')

for i in range(len(sheets)):
    so1 = co_obj1.get_sheet_by_name(name=sheets[i])
    so2 = co_obj2.get_sheet_by_name(name=sheets2[i])
    mo = wb_obj.get_sheet_by_name(name=sheets[i])

    print(sheets[i], '----', sheets2[i])

    for x in the_map[i]:
        print(x)
        le = so1[x[0]].value
        if(x[1][0] == '+'):
            pe = so2[x[1][1]].value
        if(x[1][0] == '-'):
            pe = so2[x[1][1]].value*-1
        print(le, '--------', pe)
        if(le == pe):

            so1[x[0]].fill = greenFill
            so2[x[1][1]].fill = greenFill
            print(True)
        if(le != pe):
            so1[x[0]].fill = redFill
            so2[x[1][1]].fill = redFill
            print(False)

co_obj1.save('xbrl_form.xlsx')
co_obj1.close
co_obj2.save('template.xlsx')
co_obj2.close

print(" Comparison complete: You may now review results in the template.xlsx file")
print(" Developed by virapy22(2020)")